"""
database.py — SQLite-backed storage for QDI Monitor
Tables:
  fact_table     — every parsed row from every email
  alert_history  — deduped alert lifecycle (open/resolved)
  watermark      — tracks last processed Excel row index
"""
from __future__ import annotations  # Fixes: dict | None, int | None on Python < 3.10

import sqlite3
import uuid
from datetime import datetime, timezone
from pathlib import Path

DB_PATH = 'qdi_monitor.db'


def get_conn() -> sqlite3.Connection:
    """
    Returns a connection with:
      - WAL mode  (concurrent reads + 1 writer — eliminates most lock errors)
      - 30-second busy timeout  (waits instead of crashing when DB is locked)
      - Row factory for dict-like row access
    """
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA cache_size=-64000")
    return conn


def init_db():
    """Create tables and indexes if they don't exist. Safe to call on every startup."""
    conn = None
    try:
        conn = get_conn()
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS fact_table (
            id                   INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id               INTEGER,
            email_id             TEXT,
            email_received_ts    TEXT,
            run_date             TEXT,
            job                  TEXT,
            subject              TEXT,
            table_name           TEXT,
            target_table         TEXT,
            source_count         INTEGER,
            target_count         INTEGER,
            difference_count     INTEGER,
            pct_diff             REAL,
            source_last_updated  TEXT,
            target_last_updated  TEXT,
            alert_level          TEXT DEFAULT 'NORMAL',
            inserted_at          TEXT DEFAULT (datetime('now'))
        );

        CREATE INDEX IF NOT EXISTS idx_fact_table_name ON fact_table(table_name);
        CREATE INDEX IF NOT EXISTS idx_fact_run_date   ON fact_table(run_date);
        CREATE INDEX IF NOT EXISTS idx_fact_run_id     ON fact_table(run_id);

        CREATE TABLE IF NOT EXISTS alert_history (
            alert_id                TEXT PRIMARY KEY,
            table_name              TEXT NOT NULL,
            run_id                  INTEGER,
            alert_level             TEXT,
            difference_count        INTEGER,
            pct_diff                REAL,
            previous_difference_count INTEGER,
            previous_pct_diff       REAL,
            first_seen_ts           TEXT,
            last_seen_ts            TEXT,
            status                  TEXT DEFAULT 'Open',
            reason                  TEXT,
            job                     TEXT
        );

        CREATE INDEX IF NOT EXISTS idx_alert_table  ON alert_history(table_name);
        CREATE INDEX IF NOT EXISTS idx_alert_status ON alert_history(status);

        CREATE TABLE IF NOT EXISTS watermark (
            id           INTEGER PRIMARY KEY CHECK (id = 1),
            last_row     INTEGER DEFAULT 0,
            last_run_ts  TEXT
        );

        INSERT OR IGNORE INTO watermark(id, last_row) VALUES (1, 0);
        """)
        conn.commit()
    except sqlite3.OperationalError as e:
        print(f"[database] init_db warning: {e}")
    finally:
        if conn:
            conn.close()


def get_watermark() -> int:
    conn = get_conn()
    try:
        row = conn.execute("SELECT last_row FROM watermark WHERE id=1").fetchone()
        return row['last_row'] if row else 0
    finally:
        conn.close()


def set_watermark(row_count: int):
    conn = get_conn()
    try:
        conn.execute(
            "UPDATE watermark SET last_row=?, last_run_ts=? WHERE id=1",
            (row_count, datetime.now(timezone.utc).isoformat())
        )
        conn.commit()
    finally:
        conn.close()


def insert_fact_rows(records: list[dict]):
    if not records:
        return
    conn = get_conn()
    try:
        conn.executemany("""
            INSERT INTO fact_table (
                run_id, email_id, email_received_ts, run_date, job, subject,
                table_name, target_table, source_count, target_count,
                difference_count, pct_diff, source_last_updated,
                target_last_updated, alert_level
            ) VALUES (
                :run_id, :email_id, :email_received_ts, :run_date, :job, :subject,
                :table_name, :target_table, :source_count, :target_count,
                :difference_count, :pct_diff, :source_last_updated,
                :target_last_updated, :alert_level
            )
        """, records)
        conn.commit()
    finally:
        conn.close()


def get_last_alert(table_name: str) -> dict | None:
    conn = get_conn()
    try:
        row = conn.execute("""
            SELECT * FROM alert_history
            WHERE table_name = ?
            ORDER BY last_seen_ts DESC
            LIMIT 1
        """, (table_name,)).fetchone()
        return dict(row) if row else None
    finally:
        conn.close()


def upsert_alert(table_name: str, run_id: int, alert_level: str,
                 diff_count: int, pct_diff: float,
                 prev_diff: int | None, prev_pct: float | None,
                 reason: str, job: str, ts: str):
    conn = get_conn()
    try:
        existing = conn.execute("""
            SELECT * FROM alert_history
            WHERE table_name = ? AND status = 'Open'
            ORDER BY first_seen_ts DESC LIMIT 1
        """, (table_name,)).fetchone()

        now = ts or datetime.now(timezone.utc).isoformat()

        if existing:
            conn.execute("""
                UPDATE alert_history SET
                    alert_level = ?, difference_count = ?, pct_diff = ?,
                    previous_difference_count = ?, previous_pct_diff = ?,
                    last_seen_ts = ?, run_id = ?, reason = ?
                WHERE alert_id = ?
            """, (alert_level, diff_count, pct_diff,
                  prev_diff, prev_pct, now, run_id, reason,
                  existing['alert_id']))
        else:
            conn.execute("""
                INSERT INTO alert_history (
                    alert_id, table_name, run_id, alert_level,
                    difference_count, pct_diff,
                    previous_difference_count, previous_pct_diff,
                    first_seen_ts, last_seen_ts, status, reason, job
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (str(uuid.uuid4()), table_name, run_id, alert_level,
                  diff_count, pct_diff, prev_diff, prev_pct,
                  now, now, 'Open', reason, job))

        conn.commit()
    finally:
        conn.close()


def auto_resolve_alerts():
    conn = get_conn()
    try:
        conn.execute("""
            UPDATE alert_history SET status = 'Resolved'
            WHERE status = 'Open'
            AND table_name IN (
                SELECT f.table_name FROM fact_table f
                WHERE f.run_id = (
                    SELECT MAX(run_id) FROM fact_table WHERE table_name = f.table_name
                )
                AND abs(f.pct_diff) < 1.5
            )
        """)
        conn.commit()
    finally:
        conn.close()


def load_fact_table() -> 'pd.DataFrame':
    import pandas as pd
    conn = get_conn()
    try:
        return pd.read_sql("SELECT * FROM fact_table ORDER BY email_received_ts", conn)
    finally:
        conn.close()


def load_alert_history() -> 'pd.DataFrame':
    import pandas as pd
    conn = get_conn()
    try:
        return pd.read_sql("SELECT * FROM alert_history ORDER BY last_seen_ts DESC", conn)
    finally:
        conn.close()


def get_db_stats() -> dict:
    conn = get_conn()
    try:
        stats = {}
        stats['total_fact_rows'] = conn.execute("SELECT COUNT(*) FROM fact_table").fetchone()[0]
        stats['total_alerts']    = conn.execute("SELECT COUNT(*) FROM alert_history").fetchone()[0]
        stats['open_critical']   = conn.execute("SELECT COUNT(*) FROM alert_history WHERE status='Open' AND alert_level='CRITICAL'").fetchone()[0]
        stats['open_high']       = conn.execute("SELECT COUNT(*) FROM alert_history WHERE status='Open' AND alert_level='HIGH'").fetchone()[0]
        stats['open_alerts']     = conn.execute("SELECT COUNT(*) FROM alert_history WHERE status='Open'").fetchone()[0]
        stats['resolved_alerts'] = conn.execute("SELECT COUNT(*) FROM alert_history WHERE status='Resolved'").fetchone()[0]
        stats['watermark']       = conn.execute("SELECT last_row, last_run_ts FROM watermark WHERE id=1").fetchone()
        return stats
    finally:
        conn.close()


# ── Excel Alert Export ────────────────────────────────────────────────────────
def export_alerts_to_excel(output_path: str) -> bool:
    """
    Export full alert_history to a formatted Excel file.
    Includes an 'Alerts' sheet (colour-coded by level) and a 'Summary' sheet.
    Called automatically on every pipeline run.
    Returns True on success, False on failure.
    """
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        df = load_alert_history()
        if df.empty:
            return True  # nothing to export yet

        df['exported_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        col_order = [
            'table_name', 'job', 'alert_level', 'status',
            'difference_count', 'pct_diff',
            'previous_difference_count', 'previous_pct_diff',
            'first_seen_ts', 'last_seen_ts',
            'reason', 'run_id', 'alert_id', 'exported_at'
        ]
        col_order = [c for c in col_order if c in df.columns]
        df = df[col_order]

        df.rename(columns={
            'table_name':                'Table Name',
            'job':                       'Job',
            'alert_level':               'Alert Level',
            'status':                    'Status',
            'difference_count':          'Gap (Rows)',
            'pct_diff':                  '% Diff',
            'previous_difference_count': 'Prev Gap',
            'previous_pct_diff':         'Prev % Diff',
            'first_seen_ts':             'First Seen',
            'last_seen_ts':              'Last Seen',
            'reason':                    'Reason',
            'run_id':                    'Run ID',
            'alert_id':                  'Alert ID',
            'exported_at':               'Exported At',
        }, inplace=True)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Alerts')
            wb = writer.book
            ws = writer.sheets['Alerts']

            # Header row styling
            hdr_fill  = PatternFill("solid", fgColor="070A12")
            hdr_font  = Font(bold=True, color="00D4FF", name="Calibri", size=10)
            hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin      = Border(
                bottom=Side(style="thin", color="1E2D4A"),
                right=Side(style="thin",  color="1E2D4A"),
            )
            for cell in ws[1]:
                cell.fill = hdr_fill; cell.font = hdr_font
                cell.alignment = hdr_align; cell.border = thin
            ws.row_dimensions[1].height = 28

            # Row fill colours
            fills = {
                'CRITICAL': PatternFill("solid", fgColor="3D0A12"),
                'HIGH':     PatternFill("solid", fgColor="2D1F00"),
                'NORMAL':   PatternFill("solid", fgColor="0E1320"),
                'Resolved': PatternFill("solid", fgColor="0A1A0E"),
            }
            lvl_col    = list(df.columns).index('Alert Level') + 1
            status_col = list(df.columns).index('Status') + 1

            for row_idx in range(2, ws.max_row + 1):
                lvl    = ws.cell(row=row_idx, column=lvl_col).value or ''
                status = ws.cell(row=row_idx, column=status_col).value or ''
                fill   = fills.get('Resolved' if status == 'Resolved' else lvl, fills['NORMAL'])
                fcolor = "00AA66" if status == 'Resolved' else "E2EAF8"
                for cell in ws[row_idx]:
                    cell.fill = fill
                    cell.alignment = Alignment(vertical="center")
                    cell.border = thin
                    cell.font = Font(name="Calibri", size=9, color=fcolor)

            # Auto-column widths
            for col_idx, col_cells in enumerate(ws.columns, start=1):
                width = max((len(str(c.value or '')) for c in col_cells), default=8)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 4, 55)

            ws.freeze_panes = "A2"

            # Summary sheet
            ws2 = wb.create_sheet("Summary")
            ws2['A1'] = 'QDI Alert Export Summary'
            ws2['A1'].font = Font(bold=True, size=14, color="00D4FF")
            summary = [
                ('Generated At',   datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                ('Total Alerts',   len(df)),
                ('Open Critical',  int((df['Alert Level'] == 'CRITICAL').sum())),
                ('Open High',      int((df['Alert Level'] == 'HIGH').sum())),
                ('Resolved',       int((df['Status'] == 'Resolved').sum())),
            ]
            for r, (lbl, val) in enumerate(summary, start=3):
                ws2.cell(row=r, column=1, value=lbl).font = Font(bold=True, color="6B7FA3")
                ws2.cell(row=r, column=2, value=val)
            ws2.column_dimensions['A'].width = 20
            ws2.column_dimensions['B'].width = 30
            ws2.sheet_view.tabSelected = False

        return True

    except Exception as e:
        print(f"[export_alerts_to_excel] Error: {e}")
        return False
